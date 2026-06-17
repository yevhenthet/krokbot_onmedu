#!/usr/bin/env python3
"""
Parse KROK-1 MCQ Excel file (ONMedU format) into JSON catalog.

Excel format: each question occupies 8 consecutive rows per sheet:
  Row 1: № (question number)
  Row 2: Тема (topic)
  Row 3: Текст задання (question text)
  Row 4: Правильна відповідь / Правильный ответ (correct answer → option A)
  Row 5: Option B
  Row 6: Option C
  Row 7: Option D
  Row 8: Option E

Multiple sheets are supported (e.g. one sheet per academic year).

Usage:
  pip install openpyxl
  python3 parse_excel.py --input your_file.xlsx --output mcqs.json
  python3 parse_excel.py --input your_file.xlsx --output mcqs.json --stats
"""

import json
import argparse
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl")
    raise


CORRECT_ANSWER_MARKERS = {
    'правильна відповідь', 'правильний відповідь',
    'правильный ответ', 'правильна відповідь:',
    'правильный ответ:',
}


def is_correct_answer_row(cell_value: str) -> bool:
    if not cell_value:
        return False
    return cell_value.strip().lower().rstrip(':') in CORRECT_ANSWER_MARKERS


def detect_lang(text: str) -> str:
    if not text:
        return 'uk'
    cyrillic = sum(1 for c in text if 'Ѐ' <= c <= 'ӿ')
    latin = sum(1 for c in text if c.isalpha() and c.isascii())
    return 'uk' if cyrillic >= latin else 'en'


def cell_text(cell) -> str:
    val = cell.value
    if val is None:
        return ''
    return str(val).strip()


def parse_sheet(ws, sheet_name: str, start_id: int) -> list:
    """Parse one worksheet, return list of question dicts."""
    rows = list(ws.iter_rows(values_only=False))
    questions = []
    qid = start_id
    i = 0

    while i < len(rows):
        row = rows[i]

        # Find a row whose first non-empty cell looks like a question number
        first = cell_text(row[0]) if row else ''
        # A question block starts with a numeric or short label in col A/B
        # and the next row contains a topic. We detect by checking row+3
        # for a correct-answer marker.
        if i + 7 > len(rows):
            break

        # Check if row[i+2] (3rd row of potential block) is the correct-answer marker
        row3 = rows[i + 2] if i + 2 < len(rows) else None
        col_b_row3 = cell_text(row3[1]) if row3 and len(row3) > 1 else ''

        if row3 and is_correct_answer_row(col_b_row3):
            # This is a question block: rows i..i+7
            block = rows[i:i + 8]

            def col(r_idx, c_idx=2):
                r = block[r_idx] if r_idx < len(block) else []
                return cell_text(r[c_idx]) if len(r) > c_idx else ''

            topic    = col(1)      # row 2, col C
            question = col(2)      # row 3, col C  (question text)
            ans_a    = col(3)      # row 4, col C  (correct answer = A)
            ans_b    = col(4)      # row 5
            ans_c    = col(5)      # row 6
            ans_d    = col(6)      # row 7
            ans_e    = col(7)      # row 8

            # Skip completely blank blocks
            if not question and not ans_a:
                i += 8
                continue

            # Try to get source info from col A row 1
            source = cell_text(block[0][1]) if len(block[0]) > 1 else ''

            options = {'A': ans_a or '-'}
            if ans_b: options['B'] = ans_b
            if ans_c: options['C'] = ans_c
            if ans_d: options['D'] = ans_d
            if ans_e: options['E'] = ans_e

            questions.append({
                'id':      qid,
                'sheet':   sheet_name,
                'source':  source,
                'lang':    detect_lang(question),
                'topic':   topic or None,
                'question': question,
                'options': options,
                'correct': 'A',
            })
            qid += 1
            i += 8
        else:
            i += 1

    return questions


def parse_excel(input_path: str, output_path: str, verbose: bool = False):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    all_questions = []
    qid = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        qs = parse_sheet(ws, sheet_name, qid)
        if qs:
            qid += len(qs)
            all_questions.extend(qs)
            if verbose:
                print(f"  Аркуш «{sheet_name}»: {len(qs)} питань")

    catalog = {
        'meta': {
            'source_file': Path(input_path).name,
            'total': len(all_questions),
            'sheets': wb.sheetnames,
        },
        'questions': all_questions,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    print(f"\nЗбережено {len(all_questions)} питань → {output_path}")
    return catalog


def print_stats(catalog: dict):
    from collections import Counter
    qs = catalog['questions']
    langs  = Counter(q['lang'] for q in qs)
    sheets = Counter(q['sheet'] for q in qs)
    topics = Counter((q.get('topic') or 'без теми') for q in qs)

    print(f"\n{'='*50}")
    print(f"Всього питань: {len(qs)}")
    print(f"Мова: uk={langs.get('uk',0)}, en={langs.get('en',0)}")
    print(f"\nАркушів: {len(sheets)}")
    for s, c in sorted(sheets.items()):
        print(f"  {c:3d}  {s}")
    print(f"\nТоп-15 тем:")
    for t, c in topics.most_common(15):
        print(f"  {c:3d}  {t}")

    missing_answer = [q['id'] for q in qs if not q['options'].get('A') or q['options']['A'] == '-']
    if missing_answer:
        print(f"\n⚠️  Відсутня правильна відповідь: id={missing_answer}")
    else:
        print(f"\n✅ Всі питання мають правильну відповідь")


def main():
    parser = argparse.ArgumentParser(description='Parse KROK-1 Excel → JSON')
    parser.add_argument('--input',  required=True, help='Path to .xlsx file')
    parser.add_argument('--output', default='krok1_mcqs.json', help='Output JSON file')
    parser.add_argument('--stats',  action='store_true', help='Print statistics after parsing')
    args = parser.parse_args()

    print(f"Читаю: {args.input}")
    catalog = parse_excel(args.input, args.output, verbose=True)

    if args.stats:
        print_stats(catalog)


if __name__ == '__main__':
    main()
