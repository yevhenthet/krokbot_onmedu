#!/usr/bin/env python3
"""
Parse MCQ files of ANY format using Claude AI.

Use this script when parse_excel.py fails (non-standard layout, merged cells,
PDF-converted text, Word documents, or any other irregular format).
Claude reads the raw content and extracts questions into the standard JSON catalog.

Supported input formats:
  .xlsx / .xls   — Excel (text extracted sheet by sheet)
  .txt / .md     — Plain text
  .csv           — CSV (treated as text)
  (PDF/Word)     — Convert to .txt first, then pass here

Usage:
  pip install anthropic openpyxl
  ANTHROPIC_API_KEY=sk-ant-... python3 parse_claude.py --input file.xlsx --output mcqs.json
  ANTHROPIC_API_KEY=sk-ant-... python3 parse_claude.py --input file.txt  --output mcqs.json --chunk 50
"""

import json
import argparse
import os
import sys
import time
from pathlib import Path

try:
    import anthropic
except ImportError:
    print("Run: pip install anthropic")
    sys.exit(1)

# ── Prompts ──────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Ти — парсер тестових питань. Твоє завдання — витягнути питання з наданого тексту і повернути ТІЛЬКИ валідний JSON без жодних пояснень.

Формат відповіді — масив об'єктів:
[
  {
    "topic": "назва теми або null",
    "question": "текст питання",
    "options": {
      "A": "правильна відповідь",
      "B": "варіант",
      "C": "варіант",
      "D": "варіант",
      "E": "варіант або null якщо немає"
    }
  }
]

Правила:
- Варіант A — завжди ПРАВИЛЬНА відповідь (перший варіант після позначки "Правильна відповідь" / "Правильный ответ")
- Якщо варіантів менше 5 — залиш відсутні як null
- Якщо тема не вказана — поверни null
- Повертай ТІЛЬКИ JSON-масив, без markdown, без пояснень
- Якщо у блоці тексту немає питань — поверни порожній масив []"""

USER_TEMPLATE = """Витягни всі тестові питання з наступного тексту. Правильна відповідь завжди позначена як "Правильна відповідь" або "Правильный ответ" і стає варіантом A.

ТЕКСТ:
{text}"""

# ── Text extraction ───────────────────────────────────────────────────────────

def extract_text_from_excel(path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        print("Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n=== Аркуш: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            line = '\t'.join(cells).strip()
            if line:
                parts.append(line)
    return '\n'.join(parts)


def extract_text(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext in ('.xlsx', '.xls'):
        return extract_text_from_excel(path)
    with open(path, encoding='utf-8', errors='replace') as f:
        return f.read()


# ── Chunking ──────────────────────────────────────────────────────────────────

def split_into_chunks(text: str, lines_per_chunk: int) -> list:
    lines = text.splitlines()
    chunks = []
    for i in range(0, len(lines), lines_per_chunk):
        chunk = '\n'.join(lines[i:i + lines_per_chunk]).strip()
        if chunk:
            chunks.append(chunk)
    return chunks


# ── Claude parsing ────────────────────────────────────────────────────────────

def parse_chunk(client, chunk_text: str, retries: int = 3) -> list:
    for attempt in range(retries):
        try:
            msg = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{
                    'role': 'user',
                    'content': USER_TEMPLATE.format(text=chunk_text)
                }]
            )
            raw = msg.content[0].text.strip()
            # Strip markdown code fences if present
            if raw.startswith('```'):
                raw = '\n'.join(raw.split('\n')[1:])
            if raw.endswith('```'):
                raw = '\n'.join(raw.split('\n')[:-1])
            return json.loads(raw)
        except json.JSONDecodeError as e:
            print(f"    ⚠️  JSON parse error (attempt {attempt+1}): {e}")
            if attempt < retries - 1:
                time.sleep(2)
        except Exception as e:
            print(f"    ⚠️  API error (attempt {attempt+1}): {e}")
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return []


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Parse MCQs from any format using Claude AI'
    )
    parser.add_argument('--input',  required=True, help='Input file (.xlsx, .txt, .csv)')
    parser.add_argument('--output', default='krok1_mcqs.json', help='Output JSON file')
    parser.add_argument('--chunk',  type=int, default=80,
                        help='Lines per API call (default: 80). Lower if getting errors.')
    parser.add_argument('--delay',  type=float, default=0.5,
                        help='Delay between API calls in seconds (default: 0.5)')
    args = parser.parse_args()

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print('❌ Set ANTHROPIC_API_KEY environment variable')
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    print(f'Читаю: {args.input}')
    text = extract_text(args.input)
    chunks = split_into_chunks(text, args.chunk)
    print(f'Розбито на {len(chunks)} блоків по ~{args.chunk} рядків')

    all_questions = []
    qid = 1

    for i, chunk in enumerate(chunks, 1):
        print(f'  [{i}/{len(chunks)}] Обробляю блок...', end=' ')
        results = parse_chunk(client, chunk)
        if results:
            for q in results:
                if not q.get('question'):
                    continue
                opts = q.get('options', {})
                all_questions.append({
                    'id':       qid,
                    'sheet':    Path(args.input).stem,
                    'source':   '',
                    'lang':     'uk',
                    'topic':    q.get('topic'),
                    'question': q['question'],
                    'options':  {k: v for k, v in opts.items() if v},
                    'correct':  'A',
                })
                qid += 1
            print(f'+{len(results)} питань')
        else:
            print('(порожньо)')
        time.sleep(args.delay)

    # Remove duplicates by question text
    seen = set()
    unique = []
    for q in all_questions:
        key = q['question'][:80]
        if key not in seen:
            seen.add(key)
            unique.append(q)

    # Re-number
    for i, q in enumerate(unique, 1):
        q['id'] = i

    catalog = {
        'meta': {
            'source_file': Path(args.input).name,
            'total': len(unique),
            'parser': 'claude',
        },
        'questions': unique,
    }

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    missing = [q['id'] for q in unique if not q['options'].get('A')]
    print(f'\n✅ Збережено {len(unique)} питань → {args.output}')
    if missing:
        print(f'⚠️  Без правильної відповіді: id={missing}')


if __name__ == '__main__':
    main()
